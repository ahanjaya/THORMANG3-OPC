#!/usr/bin/env python3

import os
import rospy
import rospkg
import xlsxwriter
import numpy as np 
import pandas as pd
from openpyxl.styles import Alignment
from openpyxl import load_workbook, Workbook

class Excel:
    def __init__(self, file_path):
        self.file_path = file_path
        self.create_file()
        
    def create_file(self):

        if not os.path.exists(self.file_path):
            workbook  = xlsxwriter.Workbook(self.file_path)
            worksheet = workbook.add_worksheet()

            merge_format = workbook.add_format({
                'bold': 1,
                'border': 1,
                'align': 'center',
                'valign': 'vcenter'})

            worksheet.set_column('H:J', 20) 

            worksheet.merge_range('A1:A2', 'No.', merge_format)
            worksheet.merge_range('B1:D1', 'Actual Keyboard', merge_format)
    
            worksheet.write('B2', 'X', merge_format)
            worksheet.write('C2', 'Y', merge_format)
            worksheet.write('D2', 'Theta', merge_format)
            
            worksheet.merge_range('E1:G1', 'Simulation Keyboard.', merge_format)
            worksheet.write('E2', 'X', merge_format)
            worksheet.write('F2', 'Y', merge_format)
            worksheet.write('G2', 'Theta', merge_format)
            
            worksheet.merge_range('H1:H2', 'Position Error', merge_format)
            worksheet.merge_range('I1:I2', 'Theta Error', merge_format)
            worksheet.merge_range('J1:J2', 'Mapping Theta Error', merge_format)

        else:
            rospy.loginfo('[Excel] File : {} is exist'.format(self.file_path))

    def read_data(self):
        reader = pd.read_excel(self.file_path)
        print(reader)

    def add_data(self, no, x_actual, y_actual, theta_actual, \
                       x_simula, y_simula, theta_simula, \
                       pos_err,  theta_err, map_theta_err ):

        workbook  = load_workbook(self.file_path)
        worksheet = workbook.active

        new_row = [[no, x_actual, y_actual, theta_actual, \
                        x_simula, y_simula, theta_simula, \
                        pos_err,  theta_err, map_theta_err ]]

        for data in new_row:
            worksheet.append(data)

        max_row    = worksheet.max_row
        max_column = worksheet.max_column

        for j in range(1, max_column+1):
            cell = worksheet.cell(row=max_row,column=j)
            cell.alignment = Alignment(horizontal='center') 

        # # print all value
        # for i in range(1,max_row+1):
        #     for j in range(1,max_column+1):
        #         cell = worksheet.cell(row=i,column=j)
        #         print(cell.value,end=' | ')
        #     print('\n')

        workbook.save(filename=self.file_path)

#     def run(self):
#         self.read_data()
#         self.add_data (no=1, x_actual=2, y_actual=3, theta_actual=4, \
#                              x_simula=5, y_simula=6, theta_simula=7, \
#                              pos_err=8,  theta_err=9, map_theta_err=10 )

# if __name__ == '__main__':
#     rospack   = rospkg.RosPack()
#     file_path = rospack.get_path("pioneer_main") + "/scripts/keyboard/data/"

#     excel = Excel(file_path + 'keyboard_placement.xlsx')
#     excel.run()